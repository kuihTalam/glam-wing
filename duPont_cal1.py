# -*- coding: utf-8 -*-
"""
Created on Mon Jan 30 13:10:28 2017

@author: songheng
"""

from openpyxl import load_workbook
import pandas as pd
import re
import numpy as np

baconFile = open('C:\\Users\\songheng\\Desktop\\Zach\\automate\\FFL_dupont.txt', 'r')
# 4952, 5116, 
baconContent = baconFile.read()
baconList = baconContent.splitlines()
length = float(len(baconList))

def grabRowData(item, rowNum, sheetName):
    for i in range(5, 10):
        rowRawData = (sheetName.cell(row=rowNum, column=i).value)
        if rowRawData == str('-') or rowRawData == str('n.a.') \
        or rowRawData == str('n.m.') or rowRawData == str(' '):
            data = 0
        else:
            data = float(re.sub(r'^\((.*?)\)$', r'-\1', rowRawData).replace(',',''))
        item.append(data)
        
# Loop over the xlsx files and clean up
try:
    for m in baconList:
        print(m)
        wb = load_workbook(r'C:\Users\songheng\Desktop\KLSE\Excel\%s.xlsx' % m)
    
        sheet_PL = wb.get_sheet_by_name('PL')
        sheet_BS = wb.get_sheet_by_name('BS')
        sheet_CF = wb.get_sheet_by_name('CF')
        sheet_Ratios = wb.get_sheet_by_name('Ratios')
        
        #stock_PL = {'revenue': 2,
        #            'cost_of_revenue': 4,
        #            'finance_costs': 9,
        #            'PBT': 13,
        #            'PAT': 15,}
        
        revenue = [] 
        cost_of_revenue = []
        gross_profit = []
        finance_costs = []
        PBT = []
        PAT = []
        EPS = []
        gross_div_per_share = []
        total_assets = []
        current_assets = []
        current_liabilities = []
        inventories = []
        receivable = []
        free_cash_flow = []
        PER = []
        revenue_growth = []
        net_earnings_growth = []
        ROA = []
        ROE  = []
        total_assets_turnover = []
        net_debt_to_equity = []
        interest_coverage = []
        current_ratio = []
        quality_earnings = []
        
#        grabRowData(revenue, 2, sheet_PL)
#        grabRowData(cost_of_revenue, 4, sheet_PL)
#        grabRowData(gross_profit, 6, sheet_PL)
#        grabRowData(finance_costs, 9, sheet_PL)
#        grabRowData(PBT, 13, sheet_PL)
#        grabRowData(PAT, 15, sheet_PL)
        grabRowData(EPS, 3, sheet_Ratios)
#        grabRowData(gross_div_per_share, 28, sheet_PL)
#        grabRowData(total_assets, 49, sheet_BS)
#        grabRowData(current_assets, 39, sheet_BS)
#        grabRowData(current_liabilities, 21, sheet_BS)
#        grabRowData(inventories, 45, sheet_BS)
#        grabRowData(receivable, 43, sheet_BS)
#        grabRowData(free_cash_flow, 39, sheet_CF)
#        grabRowData(PER, 10, sheet_CF)
        grabRowData(ROA, 80, sheet_Ratios)
        grabRowData(net_earnings_growth, 61, sheet_Ratios)
        grabRowData(total_assets_turnover, 91, sheet_Ratios)
        grabRowData(ROE, 82, sheet_Ratios)
        grabRowData(net_debt_to_equity, 100, sheet_Ratios)
        grabRowData(interest_coverage, 98, sheet_Ratios)
        grabRowData(current_ratio, 109, sheet_Ratios)
        grabRowData(quality_earnings, 75, sheet_Ratios)
        
        
        print('Revenue:', revenue)
        print(cost_of_revenue)
        print(gross_profit)
        print(finance_costs)
        print(PBT)
        print(PAT)
        print(EPS)
        print(gross_div_per_share)
        print(total_assets)
        print(current_assets)
        print(current_liabilities)
        print(inventories)
        print(receivable)
        print(free_cash_flow)
        print(PER)
        print(revenue_growth)
        print(net_earnings_growth)
        print(ROA)
        print(ROE)
        print(total_assets_turnover)
        print(net_debt_to_equity)
        print(interest_coverage)
        print(current_ratio)
        print(quality_earnings)

except Exception as e:
    print(e)
    pass

#gross_dividend = []
#grabRowData(gross_dividend, 28, sheet_PL)
#print(gross_dividend)
#
#PAT = []
#grabRowData(PAT, 15, sheet_PL)
#print(PAT)
#
#Long_Term_Liabilities = []
#grabRowData(Long_Term_Liabilities, 15, sheet_BS)
#print(Long_Term_Liabilities)

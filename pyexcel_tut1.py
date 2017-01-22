# -*- coding: utf-8 -*-
"""
Created on Sat Jan 21 23:56:22 2017

@author: songheng
"""


from openpyxl import load_workbook
import pandas as pd
import re
import numpy as np
wb = load_workbook(filename= '6947.xlsx')
sheet = wb.get_sheet_by_name('PL')

revenue = []
for j in range(5, 14):
    data = float(sheet.cell(row=2, column=j).value)
    revenue.append(data)

net_profit = []
for j in range(5, 14):
    raw_profit = sheet.cell(row=19, column=j).value
    prof_num = float(re.sub(r'^\((.*?)\)$', r'-\1', raw_profit).replace(',',''))
    net_profit.append(prof_num)

net_margin = [((net_profit / revenue) * 100) for revenue, net_profit in zip(revenue, net_profit)]
net_margin.reverse()
print(net_margin)
